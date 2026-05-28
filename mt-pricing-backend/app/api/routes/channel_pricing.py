"""Channel Pricing Engine — configuration and margin endpoints.

Endpoints prefix: /pricing/{channel_code}

- GET  /params                    — route + fee + scheme params
- PATCH /route-params             — update FX, freight, arancel, etc.
- PATCH /fee-params               — update commissions per channel
- GET  /margin-targets            — list family margin targets (with family_name JOIN)
- PUT  /margin-targets            — upsert margin target; clears overrides for family
- PUT  /margin-overrides/{sku}    — upsert per-SKU override
- DELETE /margin-overrides/{sku}  — remove SKU override (revert to family)
"""

from __future__ import annotations

import io as _io
import uuid
from decimal import Decimal
from typing import Annotated, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import delete, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db_session, require_permissions
from app.db.enums import CeilingBasis, FulfillmentScheme, SellingModel
from app.db.models.channel_pricing import (
    ChannelFeeParams,
    ChannelMarginOverride,
    ChannelMarginTarget,
    ChannelProductLogistics,
    ChannelSchemeParams,
    TradeRouteParams,
)
from app.db.models.channels import Channel
from app.db.models.product import Product
from app.db.models.user import User
from app.db.models.vocabularies import Family
from app.schemas.channel_pricing import (
    CatalogImportResult,
    CatalogSemaforo,
    CatalogSummaryResponse,
    ChannelFeeParamsRead,
    ChannelFeeParamsUpdate,
    ChannelSchemeParamsRead,
    MarginOverrideRead,
    MarginOverrideUpsert,
    MarginTargetRead,
    MarginTargetUpsert,
    OptimizeResponse,
    PriceResultJSON,
    ProductPriceResponse,
    ProposeSelectedRequest,
    ProposeSelectedResult,
    TradeRouteParamsRead,
    TradeRouteParamsUpdate,
)
from app.services.pricing.engine import PricingEngine
from app.services.pricing.loader import ParameterLoader
from app.services.pricing.optimizer import ChannelOptimizer
from app.services.pricing.schemas import (
    ProductLogistics,
    ProductPricingData,
)
from app.services.pricing.schemas import (
    RouteParams as _RouteParams,
)

router = APIRouter(prefix="/pricing/{channel_code}", tags=["channel-pricing"])


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


async def _resolve_channel_id(
    channel_code: str,
    session: AsyncSession,
) -> uuid.UUID:
    """Resolve channel_code → channel.id. Raises 404 if not found."""
    row = (
        await session.execute(select(Channel.id).where(Channel.code == channel_code))
    ).scalar_one_or_none()
    if row is None:
        raise HTTPException(
            status_code=404,
            detail=f"Channel '{channel_code}' not found",
        )
    return row


# ---------------------------------------------------------------------------
# GET /params — read all config for this channel
# ---------------------------------------------------------------------------


@router.get(
    "/params",
    summary="Channel pricing config (route + fees + schemes)",
    operation_id="channelPricingGetParams",
)
async def get_params(
    channel_code: str,
    _user: Annotated[User, Depends(require_permissions("prices:read"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> dict:
    """Return route + fee + scheme params for this channel."""
    channel_id = await _resolve_channel_id(channel_code, session)

    fee_row = (
        (
            await session.execute(
                select(ChannelFeeParams).where(ChannelFeeParams.channel_id == channel_id)
            )
        )
        .scalars()
        .first()
    )
    if fee_row is None:
        raise HTTPException(404, detail=f"Channel '{channel_code}' has no fee params configured")

    route_row = (
        (
            await session.execute(
                select(TradeRouteParams).where(TradeRouteParams.id == fee_row.route_id)
            )
        )
        .scalars()
        .first()
    )
    if route_row is None:
        raise HTTPException(500, detail="Trade route params missing — data integrity issue")

    scheme_rows = (
        (
            await session.execute(
                select(ChannelSchemeParams).where(ChannelSchemeParams.channel_id == channel_id)
            )
        )
        .scalars()
        .all()
    )

    total_fees_pct = float(
        fee_row.commission_pct + fee_row.vat_pct + fee_row.advertising_pct + fee_row.returns_pct
    )

    return {
        "route": TradeRouteParamsRead.model_validate(route_row).model_dump(mode="json"),
        "fees": {
            **ChannelFeeParamsRead.model_validate(fee_row).model_dump(mode="json"),
            "total_fees_pct": total_fees_pct,
        },
        "schemes": [
            ChannelSchemeParamsRead.model_validate(s).model_dump(mode="json") for s in scheme_rows
        ],
    }


# ---------------------------------------------------------------------------
# PATCH /route-params — update FX, freight, arancel, etc.
# ---------------------------------------------------------------------------


@router.patch(
    "/route-params",
    response_model=TradeRouteParamsRead,
    summary="Update trade route parameters",
    operation_id="channelPricingPatchRouteParams",
)
async def update_route_params(
    channel_code: str,
    body: TradeRouteParamsUpdate,
    _user: Annotated[User, Depends(require_permissions("prices:propose"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> TradeRouteParamsRead:
    """Update trade route parameters (FX, freight, arancel…) for this channel."""
    channel_id = await _resolve_channel_id(channel_code, session)

    fee_row = (
        (
            await session.execute(
                select(ChannelFeeParams).where(ChannelFeeParams.channel_id == channel_id)
            )
        )
        .scalars()
        .first()
    )
    if fee_row is None:
        raise HTTPException(404, detail=f"Channel '{channel_code}' has no fee params configured")

    values = body.model_dump(exclude_unset=True)
    if values:
        await session.execute(
            update(TradeRouteParams).where(TradeRouteParams.id == fee_row.route_id).values(**values)
        )
        await session.commit()

    route = (
        (
            await session.execute(
                select(TradeRouteParams).where(TradeRouteParams.id == fee_row.route_id)
            )
        )
        .scalars()
        .first()
    )
    return TradeRouteParamsRead.model_validate(route)


# ---------------------------------------------------------------------------
# PATCH /fee-params — update commissions per channel
# ---------------------------------------------------------------------------


@router.patch(
    "/fee-params",
    response_model=ChannelFeeParamsRead,
    summary="Update channel fee parameters",
    operation_id="channelPricingPatchFeeParams",
)
async def update_fee_params(
    channel_code: str,
    body: ChannelFeeParamsUpdate,
    _user: Annotated[User, Depends(require_permissions("prices:propose"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> ChannelFeeParamsRead:
    """Update channel-specific fee parameters (commission, VAT, advertising, returns…)."""
    channel_id = await _resolve_channel_id(channel_code, session)

    values = body.model_dump(exclude_unset=True)
    if values:
        await session.execute(
            update(ChannelFeeParams)
            .where(ChannelFeeParams.channel_id == channel_id)
            .values(**values)
        )
        await session.commit()

    row = (
        (
            await session.execute(
                select(ChannelFeeParams).where(ChannelFeeParams.channel_id == channel_id)
            )
        )
        .scalars()
        .first()
    )
    if row is None:
        raise HTTPException(404, detail="Channel fee params not configured")
    return ChannelFeeParamsRead.model_validate(row)


# ---------------------------------------------------------------------------
# GET /margin-targets — list with family_name JOIN
# ---------------------------------------------------------------------------


@router.get(
    "/margin-targets",
    response_model=list[MarginTargetRead],
    summary="List margin targets for channel",
    operation_id="channelPricingListMarginTargets",
)
async def list_margin_targets(
    channel_code: str,
    _user: Annotated[User, Depends(require_permissions("prices:read"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> list[MarginTargetRead]:
    """List margin targets for this channel, with family name joined in."""
    channel_id = await _resolve_channel_id(channel_code, session)

    rows = (
        await session.execute(
            select(ChannelMarginTarget, Family.name)
            .join(Family, Family.id == ChannelMarginTarget.family_id)
            .where(ChannelMarginTarget.channel_id == channel_id)
            .order_by(Family.name)
        )
    ).all()

    return [
        MarginTargetRead(
            id=r.ChannelMarginTarget.id,
            channel_id=r.ChannelMarginTarget.channel_id,
            family_id=r.ChannelMarginTarget.family_id,
            family_name=r.name,
            selling_model=SellingModel(r.ChannelMarginTarget.selling_model),
            margin_target_pct=r.ChannelMarginTarget.margin_target_pct,
        )
        for r in rows
    ]


# ---------------------------------------------------------------------------
# PUT /margin-targets — upsert + clear overrides for family
# ---------------------------------------------------------------------------


@router.put(
    "/margin-targets",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Upsert margin target (clears overrides for family)",
    operation_id="channelPricingUpsertMarginTarget",
)
async def upsert_margin_target(
    channel_code: str,
    body: MarginTargetUpsert,
    _user: Annotated[User, Depends(require_permissions("prices:propose"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> None:
    """Upsert margin target. Clears all overrides for this family+selling_model."""
    channel_id = await _resolve_channel_id(channel_code, session)

    await session.execute(
        pg_insert(ChannelMarginTarget)
        .values(
            channel_id=channel_id,
            family_id=body.family_id,
            selling_model=body.selling_model.value,
            margin_target_pct=body.margin_target_pct,
        )
        .on_conflict_do_update(
            constraint="uq_channel_margin_targets",
            set_={"margin_target_pct": body.margin_target_pct},
        )
    )
    # Clear overrides for products in this family (Pricing Desk behavior)
    await session.execute(
        delete(ChannelMarginOverride).where(
            ChannelMarginOverride.channel_id == channel_id,
            ChannelMarginOverride.selling_model == body.selling_model.value,
            ChannelMarginOverride.product_sku.in_(
                select(Product.sku).where(Product.family_id == body.family_id)
            ),
        )
    )
    await session.commit()


# ---------------------------------------------------------------------------
# PUT /margin-overrides/{sku} — upsert per-SKU override
# ---------------------------------------------------------------------------


@router.put(
    "/margin-overrides/{sku}",
    response_model=MarginOverrideRead,
    summary="Upsert per-SKU margin override",
    operation_id="channelPricingUpsertMarginOverride",
)
async def upsert_margin_override(
    channel_code: str,
    sku: str,
    body: MarginOverrideUpsert,
    _user: Annotated[User, Depends(require_permissions("prices:propose"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> MarginOverrideRead:
    """Upsert per-SKU margin override."""
    channel_id = await _resolve_channel_id(channel_code, session)

    await session.execute(
        pg_insert(ChannelMarginOverride)
        .values(
            product_sku=sku,
            channel_id=channel_id,
            selling_model=body.selling_model.value,
            margin_override_pct=body.margin_override_pct,
            reason=body.reason,
        )
        .on_conflict_do_update(
            constraint="uq_channel_margin_overrides",
            set_={
                "margin_override_pct": body.margin_override_pct,
                "reason": body.reason,
            },
        )
    )
    await session.commit()

    row = (
        (
            await session.execute(
                select(ChannelMarginOverride).where(
                    ChannelMarginOverride.product_sku == sku,
                    ChannelMarginOverride.channel_id == channel_id,
                    ChannelMarginOverride.selling_model == body.selling_model.value,
                )
            )
        )
        .scalars()
        .first()
    )
    return MarginOverrideRead.model_validate(row)


# ---------------------------------------------------------------------------
# DELETE /margin-overrides/{sku} — remove SKU override
# ---------------------------------------------------------------------------


@router.delete(
    "/margin-overrides/{sku}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Remove per-SKU margin override",
    operation_id="channelPricingDeleteMarginOverride",
)
async def delete_margin_override(
    channel_code: str,
    sku: str,
    _user: Annotated[User, Depends(require_permissions("prices:propose"))],
    session: Annotated[AsyncSession, Depends(get_db_session)],
    selling_model: SellingModel = Query(default=SellingModel.B2C),
) -> None:
    """Remove a SKU override — product reverts to family margin target."""
    channel_id = await _resolve_channel_id(channel_code, session)

    await session.execute(
        delete(ChannelMarginOverride).where(
            ChannelMarginOverride.product_sku == sku,
            ChannelMarginOverride.channel_id == channel_id,
            ChannelMarginOverride.selling_model == selling_model.value,
        )
    )
    await session.commit()


# ---------------------------------------------------------------------------
# Calculation helpers
# ---------------------------------------------------------------------------


def _price_result_to_dict(r) -> dict:
    """Serialize PriceResult to JSON-friendly dict. Infinity → None."""
    inf = Decimal("Infinity")
    return {
        "sku": r.sku,
        "selling_model": r.selling_model.value,
        "fulfillment_scheme": r.fulfillment_scheme.value,
        "scheme_label": r.scheme_label,
        "margin_pct": float(r.margin_pct),
        "cost_op_aed": float(r.cost_op_aed),
        "selling_price_aed": (float(r.selling_price_aed) if r.selling_price_aed != inf else None),
        # ceiling_aed=null means EITHER:
        # - Decimal("Infinity") (MARGIN_FLOOR basis: no PVP catalog reference)
        # - Decimal("0") (infeasible result placeholder)
        # Frontend should use `signal` and `is_publishable` as canonical indicators.
        "ceiling_aed": (float(r.ceiling_aed) if r.ceiling_aed not in (inf, Decimal("0")) else None),
        "benefit_per_unit_aed": float(r.benefit_per_unit_aed),
        "roi_pct": float(r.roi_pct),
        "margin_to_ceiling_pct": float(r.margin_to_ceiling_pct),
        "is_publishable": r.is_publishable,
        "signal": r.signal,
    }


# ---------------------------------------------------------------------------
# GET /product/{sku} — single-SKU price calculation
# ---------------------------------------------------------------------------


@router.get(
    "/product/{sku}",
    response_model=ProductPriceResponse,
    operation_id="getProductPrice",
    dependencies=[Depends(require_permissions("prices:read"))],
)
async def get_product_price(
    channel_code: str,
    sku: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    selling_model: SellingModel = Query(default=SellingModel.B2C),
    margin_pct: Optional[float] = None,
) -> ProductPriceResponse:
    """Calculate price for one SKU across all schemes + best."""
    channel_id = await _resolve_channel_id(channel_code, session)
    loader = ParameterLoader(session)
    route, fees, schemes = await loader.load_route_and_fees(channel_id)
    products = await loader.load_product_data(channel_id, skus=[sku])
    if not products:
        raise HTTPException(
            status_code=404,
            detail=f"SKU '{sku}' not found or has no logistics data",
        )

    product = products[0]
    effective_margins = await loader.load_effective_margins(channel_id, selling_model, [sku])
    m = (
        Decimal(str(margin_pct))
        if margin_pct is not None
        else effective_margins.get(sku, Decimal("12"))
    )

    compute = (
        PricingEngine.compute_b2c
        if selling_model == SellingModel.B2C
        else PricingEngine.compute_b2b
    )
    results = [compute(product, route, fees, s, m) for s in schemes if s.is_available]

    if selling_model == SellingModel.B2C:
        best = ChannelOptimizer.best_scheme_b2c(product, route, fees, schemes, m)
    else:
        best = ChannelOptimizer.best_scheme_b2b(product, route, fees, schemes, m)

    return ProductPriceResponse(
        sku=sku,
        effective_margin_pct=float(m),
        best_scheme=PriceResultJSON(**_price_result_to_dict(best)) if best else None,
        all_schemes=[PriceResultJSON(**_price_result_to_dict(r)) for r in results],
    )


# ---------------------------------------------------------------------------
# GET /catalog — full catalog summary with semáforo + filters
# ---------------------------------------------------------------------------


@router.get(
    "/catalog",
    response_model=CatalogSummaryResponse,
    operation_id="getCatalogSummary",
    dependencies=[Depends(require_permissions("prices:read"))],
)
async def get_catalog_summary(
    channel_code: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    selling_model: SellingModel = Query(default=SellingModel.B2C),
    family_id: Optional[uuid.UUID] = None,
    signal: Optional[str] = None,
) -> CatalogSummaryResponse:
    """Return price analysis for the full catalog with semáforo summary."""
    channel_id = await _resolve_channel_id(channel_code, session)
    loader = ParameterLoader(session)
    route, fees, schemes = await loader.load_route_and_fees(channel_id)
    products = await loader.load_product_data(channel_id)

    if family_id:
        products = [p for p in products if p.family_id == str(family_id)]

    skus = [p.sku for p in products]
    margins = await loader.load_effective_margins(channel_id, selling_model, skus)

    if selling_model == SellingModel.B2C:
        results = ChannelOptimizer.optimize_catalog_b2c(products, route, fees, schemes, margins)
    else:
        results = ChannelOptimizer.optimize_catalog_b2b(products, route, fees, schemes, margins)

    if signal:
        results = [r for r in results if r.signal == signal.upper()]

    rows = [PriceResultJSON(**_price_result_to_dict(r)) for r in results]
    publishable = sum(1 for r in results if r.is_publishable)
    in_loss = sum(1 for r in results if r.signal == "PÉRDIDA")

    return CatalogSummaryResponse(
        semaforo=CatalogSemaforo(
            total=len(results),
            publishable=publishable,
            blocked=len(results) - publishable,
            in_loss=in_loss,
            by_scheme={
                scheme.value: sum(1 for r in results if r.fulfillment_scheme == scheme)
                for scheme in FulfillmentScheme
            },
        ),
        rows=rows,
    )


# ---------------------------------------------------------------------------
# POST /optimize — optimization preview (does NOT persist)
# ---------------------------------------------------------------------------


@router.post(
    "/optimize",
    response_model=OptimizeResponse,
    operation_id="optimizeCatalog",
    dependencies=[Depends(require_permissions("prices:read"))],
)
async def optimize_catalog(
    channel_code: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    selling_model: SellingModel = Query(default=SellingModel.B2C),
) -> OptimizeResponse:
    """Preview the best scheme+margin per product. Does NOT persist.

    PERFORMANCE: CPU-bound. For catalogs >50 SKUs, consider a Celery task.
    """
    channel_id = await _resolve_channel_id(channel_code, session)
    loader = ParameterLoader(session)
    route, fees, schemes = await loader.load_route_and_fees(channel_id)
    products = await loader.load_product_data(channel_id)

    if selling_model == SellingModel.B2C:
        results = ChannelOptimizer.full_optimize_catalog_b2c(products, route, fees, schemes)
    else:
        results = ChannelOptimizer.full_optimize_catalog_b2b(products, route, fees, schemes)

    return OptimizeResponse(results=[PriceResultJSON(**_price_result_to_dict(r)) for r in results])


# ---------------------------------------------------------------------------
# POST /optimize/apply — persist optimization as overrides
# ---------------------------------------------------------------------------


@router.post(
    "/optimize/apply",
    operation_id="applyOptimization",
    dependencies=[Depends(require_permissions("prices:propose"))],
    status_code=status.HTTP_204_NO_CONTENT,
)
async def apply_optimization(
    channel_code: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    selling_model: SellingModel = Query(default=SellingModel.B2C),
) -> None:
    """Persist optimization result as per-SKU margin overrides."""
    channel_id = await _resolve_channel_id(channel_code, session)
    loader = ParameterLoader(session)
    route, fees, schemes = await loader.load_route_and_fees(channel_id)
    products = await loader.load_product_data(channel_id)

    if selling_model == SellingModel.B2C:
        results = ChannelOptimizer.full_optimize_catalog_b2c(products, route, fees, schemes)
    else:
        results = ChannelOptimizer.full_optimize_catalog_b2b(products, route, fees, schemes)

    for r in results:
        await session.execute(
            pg_insert(ChannelMarginOverride)
            .values(
                product_sku=r.sku,
                channel_id=channel_id,
                selling_model=selling_model.value,
                margin_override_pct=r.margin_pct,
                reason="auto-optimized",
            )
            .on_conflict_do_update(
                constraint="uq_channel_margin_overrides",
                set_={
                    "margin_override_pct": r.margin_pct,
                    "reason": "auto-optimized",
                },
            )
        )
    await session.commit()


# ---------------------------------------------------------------------------
# POST /prices/propose-selected — propose prices for selected SKUs
# ---------------------------------------------------------------------------


@router.post(
    "/prices/propose-selected",
    operation_id="proposePricesSelected",
    response_model=ProposeSelectedResult,
    dependencies=[Depends(require_permissions("prices:propose"))],
)
async def propose_prices_selected(
    channel_code: str,
    body: ProposeSelectedRequest,
    session: Annotated[AsyncSession, Depends(get_db_session)],
) -> ProposeSelectedResult:
    """Propose prices for selected SKUs into the approval flow.

    Runs the pricing engine for each SKU and inserts rows into `prices`
    with status='pending_review'. SKUs not found in channel logistics are
    skipped; infeasible prices are reported as errors.

    proposed_by is stored as NULL until get_current_user is threaded through.
    # TODO: thread current user UUID once auth dependency is available here.
    """
    from app.services.pricing.price_proposer import PriceProposer

    channel_id = await _resolve_channel_id(channel_code, session)
    proposer = PriceProposer(session)
    return await proposer.propose(
        channel_id=channel_id,
        skus=body.skus,
        selling_model=body.selling_model,
        proposed_by=None,  # TODO: replace with current user UUID once threaded
        notes=body.notes,
    )


# ── Excel import endpoints ───────────────────────────────────────────


@router.post(
    "/catalog/import",
    operation_id="importCatalog",
    response_model=CatalogImportResult,
    dependencies=[Depends(require_permissions("prices:propose"))],
)
async def import_catalog(
    channel_code: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    file: UploadFile = File(...),
    confirm: bool = False,
) -> CatalogImportResult:
    """Import MT catalog Excel.

    Required columns: sku, pe_eur, pvp_eur, uds_caja, peso_kg.
    Optional: ceiling_basis (default catalog_pvp).

    Pass confirm=true to persist. Without confirm, returns preview with
    calculated ceiling prices for the first 20 valid rows.
    """
    channel_id = await _resolve_channel_id(channel_code, session)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, detail=f"Cannot read Excel file: {exc}")
    ws = wb.active

    headers_row = next(ws.iter_rows(max_row=1), None)
    if headers_row is None:
        raise HTTPException(400, detail="Empty Excel file")
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in headers_row]
    required = {"sku", "pe_eur", "pvp_eur", "uds_caja", "peso_kg"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(400, detail=f"Missing required columns: {sorted(missing)}")

    idx = {h: i for i, h in enumerate(headers)}
    errors: list[dict] = []
    valid_rows: list[dict] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        def cell(col_name: str, _row=row):
            return _row[idx[col_name]] if col_name in idx and idx[col_name] < len(_row) else None

        sku = str(cell("sku") or "").strip()
        if not sku:
            continue
        try:
            pe = Decimal(str(cell("pe_eur")))
            pvp = Decimal(str(cell("pvp_eur")))
            uds = int(cell("uds_caja") or 1)
            peso_raw = cell("peso_kg")
            peso = Decimal(str(peso_raw)) if peso_raw is not None else None
            cb_raw = str(cell("ceiling_basis") or "catalog_pvp").strip()
            try:
                cb = CeilingBasis(cb_raw)
            except ValueError:
                cb = CeilingBasis.CATALOG_PVP

            if pe <= 0:
                raise ValueError("pe_eur must be > 0")
            if pvp <= 0:
                raise ValueError("pvp_eur must be > 0")
            if uds < 1:
                raise ValueError("uds_caja must be >= 1")

            valid_rows.append(
                {
                    "sku": sku,
                    "pe_eur": pe,
                    "catalog_pvp_eur": pvp,
                    "units_per_box": uds,
                    "weight": peso,
                    "ceiling_basis": cb,
                }
            )
        except Exception as e:
            errors.append({"row": row_num, "sku": sku, "error": str(e)})

    # Build ceiling preview using current route params
    fee_row = (
        (
            await session.execute(
                select(ChannelFeeParams).where(ChannelFeeParams.channel_id == channel_id)
            )
        )
        .scalars()
        .first()
    )
    route_row = (
        (
            await session.execute(
                select(TradeRouteParams).where(TradeRouteParams.id == fee_row.route_id)
            )
        )
        .scalars()
        .first()
        if fee_row
        else None
    )

    ceiling_preview: list[dict] = []
    if route_row:
        route_dc = _RouteParams(
            fx_rate=route_row.fx_rate,
            fx_buffer_pct=route_row.fx_buffer_pct,
            freight_rate_per_kg=route_row.freight_rate_per_kg,
            freight_min_aed=route_row.freight_min_aed,
            import_tariff_pct=route_row.import_tariff_pct,
            local_warehouse_pct=route_row.local_warehouse_pct,
            handling_pct=route_row.handling_pct,
        )
        for r in valid_rows[:20]:
            dummy_logistics = ProductLogistics(
                inbound_fee_aed=Decimal("0"),
                storage_fee_aed=Decimal("0"),
                fulfillment_fee_aed=Decimal("0"),
                default_scheme=FulfillmentScheme.CANAL_FULL,
            )
            dummy_product = ProductPricingData(
                sku=r["sku"],
                family_id="preview",
                pe_eur=r["pe_eur"],
                catalog_pvp_eur=r["catalog_pvp_eur"],
                units_per_box=r["units_per_box"],
                weight_kg=r.get("weight") or Decimal("0"),
                b2c_labeling_aed=Decimal("0"),
                ceiling_basis=r["ceiling_basis"],
                logistics=dummy_logistics,
            )
            c_b2c = PricingEngine._ceiling_b2c(dummy_product, route_dc)
            c_b2b = PricingEngine._ceiling_b2b(dummy_product, route_dc)
            ceiling_preview.append(
                {
                    "sku": r["sku"],
                    "ceiling_b2c_aed": (float(c_b2c) if c_b2c != Decimal("Infinity") else None),
                    "ceiling_b2b_aed": (float(c_b2b) if c_b2b != Decimal("Infinity") else None),
                }
            )

    upserted = 0
    if confirm:
        for r in valid_rows:
            values: dict = {
                "pe_eur": r["pe_eur"],
                "catalog_pvp_eur": r["catalog_pvp_eur"],
                "units_per_box": r["units_per_box"],
                "ceiling_basis": r["ceiling_basis"].value,
            }
            if r.get("weight") is not None:
                values["weight"] = r["weight"]
            result = await session.execute(
                update(Product).where(Product.sku == r["sku"]).values(**values)
            )
            if result.rowcount > 0:
                upserted += 1
            else:
                errors.append(
                    {
                        "row": None,
                        "sku": r["sku"],
                        "error": "SKU not found in products table",
                    }
                )
        await session.commit()

    return CatalogImportResult(
        total_rows=len(valid_rows) + len(errors),
        upserted=upserted,
        errors=errors,
        ceiling_preview=ceiling_preview,
    )


@router.post(
    "/logistics/import",
    operation_id="importLogistics",
    dependencies=[Depends(require_permissions("prices:propose"))],
)
async def import_logistics(
    channel_code: str,
    session: Annotated[AsyncSession, Depends(get_db_session)],
    file: UploadFile = File(...),
    confirm: bool = False,
) -> dict:
    """Import logistics fees Excel (inbound, storage, fulfillment per SKU)."""
    channel_id = await _resolve_channel_id(channel_code, session)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, detail=f"Cannot read Excel file: {exc}")
    ws = wb.active
    headers_row = next(ws.iter_rows(max_row=1), None)
    if headers_row is None:
        raise HTTPException(400, detail="Empty Excel file")
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in headers_row]
    required = {"sku", "inbound_fee_aed", "storage_fee_aed", "fulfillment_fee_aed"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(400, detail=f"Missing columns: {sorted(missing)}")

    idx = {h: i for i, h in enumerate(headers)}
    errors: list[dict] = []
    upserted = 0
    total = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        def cell(col_name: str, _row=row):
            return _row[idx[col_name]] if col_name in idx and idx[col_name] < len(_row) else None

        sku = str(cell("sku") or "").strip()
        if not sku:
            continue
        total += 1
        try:
            scheme_raw = str(cell("default_scheme") or "canal_full").strip()
            try:
                scheme = FulfillmentScheme(scheme_raw)
            except ValueError:
                scheme = FulfillmentScheme.CANAL_FULL
            values = {
                "product_sku": sku,
                "channel_id": channel_id,
                "inbound_fee_aed": Decimal(str(cell("inbound_fee_aed") or 0)),
                "storage_fee_aed": Decimal(str(cell("storage_fee_aed") or 0)),
                "fulfillment_fee_aed": Decimal(str(cell("fulfillment_fee_aed") or 0)),
                "default_scheme": scheme.value,
            }
            if confirm:
                await session.execute(
                    pg_insert(ChannelProductLogistics)
                    .values(**values)
                    .on_conflict_do_update(
                        constraint="uq_channel_product_logistics",
                        set_={
                            k: v
                            for k, v in values.items()
                            if k not in ("product_sku", "channel_id")
                        },
                    )
                )
                upserted += 1
        except Exception as e:
            errors.append({"row": row_num, "sku": sku, "error": str(e)})

    if confirm:
        await session.commit()

    return {"total_rows": total, "upserted": upserted, "errors": errors}


__all__ = ["router"]
