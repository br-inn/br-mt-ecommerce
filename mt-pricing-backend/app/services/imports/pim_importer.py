"""PimImporter — orquestador batch async del PIM completo (US-1A-06-01).

Diseñado para correr en Celery worker:
- Recibe ``run_id`` de un ``ImportRun`` ya persistido en estado ``queued``.
- Lee xlsx desde filesystem o Storage (path se resuelve por el caller, aquí
  se asume un path local accesible por el worker — los uploads via API se
  descargan a /tmp antes de invocar).
- Itera con openpyxl ``read_only=True`` (no carga 5k filas en RAM).
- Por cada fila: cast → upsert en `products` por SKU → audit event.
- Commit periódico cada 100 filas + flush incremental del ImportRun.
- Errores por celda no abortan el run; se acumulan en ``ImportRun.errors``
  cap a 100 entradas (display).

Idempotencia:
- UPSERT por ``products.sku`` (PK natural). Re-correr el mismo PIM no
  duplica filas. Campos con ``manual_locked_fields`` NO se sobrescriben
  (alineado con el wizard sincrono — la lógica vive en `_apply_to_product`).
"""

from __future__ import annotations

import asyncio
import logging
import os
import tempfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.db.models.import_run import ImportRun
from app.db.models.product import ProductTranslation
from app.repositories.audit import AuditRepository
from app.repositories.product import ProductRepository
from app.services.importer.column_mapper import EXPECTED_HEADERS
from app.services.imports.division_assignment import assign_divisions
from app.services.imports.pim_row_mapper import map_pim_row_to_product

logger = logging.getLogger(__name__)

#: Cap de errores serializados en ImportRun.errors (display).
MAX_ERRORS_LOGGED: int = 100

#: Frecuencia de commit periódico (filas).
COMMIT_EVERY_N_ROWS: int = 100


class PimImporter:
    """Importer batch del PIM completo.xlsx.

    Args:
        session: AsyncSession con autocommit OFF — el importer hace commits
            periódicos cada 100 filas via ``session.commit()`` directo.
        source_path: path filesystem del xlsx (worker tiene el volumen montado).
        run_id: UUID del ImportRun pre-creado (estado ``queued``).
        actor_id: UUID del user que disparó el run (para audit). None permitido
            sólo en endpoint de fixture dev (run_pim_from_fixture).
    """

    def __init__(
        self,
        session: AsyncSession,
        source_path: str | Path,
        run_id: UUID | str,
        actor_id: UUID | None = None,
        storage_bucket: str | None = None,
    ) -> None:
        self.session = session
        self.source_path = Path(source_path)
        self.run_id = run_id if isinstance(run_id, UUID) else UUID(str(run_id))
        self.actor_id = actor_id
        self._run: ImportRun | None = None
        self._repo = ProductRepository(session)
        self._audit = AuditRepository(session)
        # If set, download the xlsx from Supabase Storage when source_path is absent locally.
        self._storage_bucket: str | None = storage_bucket
        self._tmp_path: str | None = None
        # Stage 3 Wave 11 — division mapping. Resuelto lazy en run().
        self._division_codes: list[str] = []
        self._division_code_cache: dict[str, UUID | None] = {}

    # ------------------------------------------------------------------ run
    async def run(self) -> ImportRun:
        """Ejecuta el import. NO lanza excepciones por filas malas — sólo por
        errores fatales (archivo no encontrado, header mismatch, DB caída).
        """
        self._run = await self.session.get(ImportRun, self.run_id)
        if self._run is None:
            raise RuntimeError(f"ImportRun {self.run_id} no existe.")

        # Stage 3 Wave 11 — resolver division_codes:
        #   1) summary.division_codes del run (preferido — TI puede setearlo
        #      por run via el endpoint de upload o admin update).
        #   2) settings.PIM_DEFAULT_DIVISIONS (fallback global).
        run_summary = self._run.summary or {}
        summary_codes = run_summary.get("division_codes")
        if isinstance(summary_codes, list) and summary_codes:
            self._division_codes = [str(c) for c in summary_codes if c]
        else:
            self._division_codes = list(settings.PIM_DEFAULT_DIVISIONS or [])
        if self._division_codes:
            logger.info(
                "PimImporter run_id=%s aplicará divisions=%s a cada SKU.",
                self.run_id,
                self._division_codes,
            )

        if not self.source_path.exists():
            if self._storage_bucket:
                await self._download_from_storage()
            else:
                await self._mark_failed(f"Archivo no encontrado: {self.source_path}")
                raise FileNotFoundError(self.source_path)

        # Load workbook streaming. data_only=True para evaluar formulas pre-cache.
        from openpyxl import load_workbook

        try:
            wb = load_workbook(str(self.source_path), read_only=True, data_only=True)
        except Exception as exc:
            await self._mark_failed(f"openpyxl load failed: {exc}")
            raise

        try:
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)

            # Header validation — abortar si no coincide con la spec.
            try:
                header = next(rows_iter)
            except StopIteration:
                await self._mark_failed("Archivo vacío (sin header).")
                return self._run

            header_errors = self._collect_header_errors(header)
            if header_errors:
                await self._mark_failed("Header mismatch: " + "; ".join(header_errors[:5]))
                return self._run

            # Marca run como running.
            self._run.status = "running"
            self._run.started_at = datetime.now(tz=UTC)
            await self.session.commit()

            inserted = 0
            updated = 0
            skipped = 0
            errors: list[dict[str, Any]] = []
            error_rows = 0
            row_idx = 1  # 1 = primera fila de datos (post-header)

            for row in rows_iter:
                # Skip filas totalmente vacias (openpyxl emite tail rows None).
                if all(v is None or v == "" for v in row):
                    skipped += 1
                    row_idx += 1
                    continue

                # Savepoint per fila — una mala no contamina el resto.
                try:
                    async with self.session.begin_nested():
                        action = await self._process_row(row, row_idx)
                    if action == "inserted":
                        inserted += 1
                    elif action == "updated":
                        updated += 1
                    else:
                        skipped += 1
                except Exception as exc:
                    error_rows += 1
                    if len(errors) < MAX_ERRORS_LOGGED:
                        errors.append(
                            {
                                "row": row_idx + 1,  # +1 porque header es row 1
                                "error": str(exc)[:200],
                            }
                        )
                    logger.warning("PimImporter row %d failed: %s", row_idx + 1, exc)

                # Commit periódico — flush al ImportRun + a products.
                if row_idx % COMMIT_EVERY_N_ROWS == 0:
                    self._run.inserted_rows = inserted
                    self._run.updated_rows = updated
                    self._run.skipped_rows = skipped
                    self._run.error_rows = error_rows
                    self._run.errors = errors
                    self._run.total_rows = row_idx
                    try:
                        await self.session.commit()
                    except Exception:
                        logger.exception(
                            "PimImporter periodic commit failed at row %d",
                            row_idx,
                        )
                        try:
                            await self.session.rollback()
                        except Exception:
                            pass
                row_idx += 1
        finally:
            wb.close()

        # Final flush.
        self._run.total_rows = row_idx - 1
        self._run.inserted_rows = inserted
        self._run.updated_rows = updated
        self._run.skipped_rows = skipped
        self._run.error_rows = error_rows
        self._run.errors = errors
        self._run.summary = {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "errors": error_rows,
            "max_errors_logged": MAX_ERRORS_LOGGED,
        }
        self._run.finished_at = datetime.now(tz=UTC)
        self._run.status = "completed" if error_rows == 0 else "completed_with_errors"
        await self.session.commit()
        logger.info(
            "PimImporter completed run_id=%s inserted=%d updated=%d errors=%d",
            self.run_id,
            inserted,
            updated,
            error_rows,
        )
        self._cleanup_tmp()
        return self._run

    async def _download_from_storage(self) -> None:
        """Download xlsx from Supabase Storage to a local temp file.

        Updates ``self.source_path`` to the temp file path on success.
        Stores path in ``self._tmp_path`` for cleanup.
        """
        from app.core.supabase import get_supabase_admin

        bucket_path = str(self.source_path)
        logger.info(
            "PimImporter: descargando %s/%s desde Supabase Storage…",
            self._storage_bucket,
            bucket_path,
        )
        try:
            loop = asyncio.get_running_loop()
            sb = get_supabase_admin()
            data: bytes = await loop.run_in_executor(
                None,
                lambda: sb.storage.from_(self._storage_bucket).download(bucket_path),
            )
        except Exception as exc:
            await self._mark_failed(f"Error descargando de Storage {self._storage_bucket}/{bucket_path}: {exc}")
            raise FileNotFoundError(
                f"Storage:{self._storage_bucket}/{bucket_path}"
            ) from exc

        fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="pim_import_")
        with os.fdopen(fd, "wb") as fh:
            fh.write(data)
        self._tmp_path = tmp
        self.source_path = Path(tmp)
        logger.info(
            "PimImporter: descargados %d bytes → %s",
            len(data),
            tmp,
        )

    def _cleanup_tmp(self) -> None:
        """Remove temp file created by _download_from_storage, if any."""
        if self._tmp_path:
            try:
                os.unlink(self._tmp_path)
            except OSError:
                pass
            self._tmp_path = None

    # --------------------------------------------------------- header check
    @staticmethod
    def _collect_header_errors(header: tuple[Any, ...]) -> list[str]:
        """Devuelve lista de mismatches contra EXPECTED_HEADERS. Vacia → OK."""
        errors: list[str] = []
        if len(header) < len(EXPECTED_HEADERS):
            errors.append(f"Header con {len(header)} columnas; esperadas {len(EXPECTED_HEADERS)}.")
            return errors
        for i, expected in enumerate(EXPECTED_HEADERS):
            actual = header[i]
            actual_str = (str(actual) if actual is not None else "").strip()
            if actual_str != expected:
                errors.append(f"col {i}: esperado {expected!r}, recibido {actual_str!r}")
        return errors

    async def _mark_failed(self, reason: str) -> None:
        """Marca el run como failed + persiste la razón en errors."""
        self._cleanup_tmp()
        if self._run is None:
            return
        self._run.status = "failed"
        self._run.finished_at = datetime.now(tz=UTC)
        existing = list(self._run.errors or [])
        existing.append({"row": 0, "error": reason[:200]})
        self._run.errors = existing[:MAX_ERRORS_LOGGED]
        try:
            await self.session.commit()
        except Exception:
            logger.exception("PimImporter _mark_failed commit failed")
            await self.session.rollback()

    # --------------------------------------------------------- row process
    async def _upsert_en_translation(self, sku: str, name_en: str) -> None:
        """Inserta o actualiza la traducción en='en' para el SKU dado."""
        stmt = (
            pg_insert(ProductTranslation)
            .values(sku=sku, lang="en", name=name_en, status="draft")
            .on_conflict_do_update(
                index_elements=["sku", "lang"],
                set_={
                    "name": name_en,
                    "status": "draft",
                    "updated_at": text("now()"),
                },
            )
        )
        await self.session.execute(stmt)

    async def _process_row(self, row: tuple[Any, ...], row_idx: int) -> str:
        """Procesa una fila: cast → upsert. Devuelve 'inserted'|'updated'|'skipped'."""
        payload = map_pim_row_to_product(row)
        cast_errors = payload.pop("_row_errors", None)
        # Limpia keys que no son columnas reales del modelo Product.
        payload.pop("_row_errors", None)

        # name_en es hybrid_property read-only en Product; hay que persistirlo
        # como ProductTranslation(lang='en'), no como campo escalar.
        name_en: str | None = payload.pop("name_en", None)
        # active es hybrid_property read-only (lifecycle_status == 'active').
        # INSERT usa server_default 'active'; UPDATE no cambia lifecycle desde PIM.
        payload.pop("active", None)

        sku = payload["sku"]
        existing = await self._repo.get_by_sku(sku)

        if existing is None:
            # INSERT — campos default vienen del payload (brand, family, etc.).
            if self.actor_id is not None:
                payload["created_by"] = self.actor_id
                payload["updated_by"] = self.actor_id
            await self._repo.create(**payload)
            if name_en:
                await self._upsert_en_translation(sku, name_en)
            await self._audit_event(
                action="product.imported.created",
                sku=sku,
                payload_diff={
                    "_import_run_id": str(self.run_id),
                    "cast_errors": cast_errors or [],
                },
                after=_safe_repr(payload),  # Decimal/datetime → str para JSONB.
            )
            await self._maybe_assign_divisions(sku)
            return "inserted"

        # UPDATE — solo campos no locked. Alineado con applier sincrono.
        locked = set(existing.manual_locked_fields or [])
        changed: dict[str, Any] = {}
        for field, new_value in payload.items():
            if field in locked:
                continue
            if field in {"sku", "internal_id", "created_at", "created_by"}:
                continue
            current = getattr(existing, field, None)
            # Comparacion shallow — JSONB se compara por igualdad de dicts.
            if current != new_value:
                setattr(existing, field, new_value)
                changed[field] = {"from": _safe_repr(current), "to": _safe_repr(new_value)}

        if name_en and "translations.en" not in locked:
            await self._upsert_en_translation(sku, name_en)

        if not changed:
            # Aún sin cambios en el producto, las divisiones pueden faltar
            # (e.g. backfill de un PIM previo). Idempotente, barato.
            await self._maybe_assign_divisions(sku)
            return "skipped"

        if self.actor_id is not None:
            existing.updated_by = self.actor_id
        existing.updated_at = datetime.now(tz=UTC)
        await self.session.flush()
        await self._audit_event(
            action="product.imported.updated",
            sku=sku,
            payload_diff={
                "_import_run_id": str(self.run_id),
                "diff": changed,
                "cast_errors": cast_errors or [],
            },
        )
        await self._maybe_assign_divisions(sku)
        return "updated"

    async def _maybe_assign_divisions(self, sku: str) -> None:
        """Asigna divisiones (Stage 3 Wave 11) si el run las trae.

        No-op si ``_division_codes`` está vacío. Tolerante: un fallo aquí no
        debe matar el upsert (ya hecho). Por consistencia, lo loggeamos pero
        propagamos para que el savepoint del row se rollbackee — así la fila
        cae a errors y se re-procesa en el próximo run, mejor que dejar
        productos sin division asignada.
        """
        if not self._division_codes:
            return
        try:
            await assign_divisions(
                self.session,
                sku,
                self._division_codes,
                code_id_cache=self._division_code_cache,
            )
        except Exception:
            logger.exception(
                "PimImporter assign_divisions failed sku=%s codes=%s",
                sku,
                self._division_codes,
            )
            raise

    async def _audit_event(
        self,
        *,
        action: str,
        sku: str,
        payload_diff: dict[str, Any],
        after: dict[str, Any] | None = None,
    ) -> None:
        """Audit event tolerante — un fallo aqui no debe matar la fila."""
        try:
            await self._audit.record(
                entity_type="product",
                entity_id=sku,
                action=action,
                actor_id=self.actor_id,
                payload_diff=payload_diff,
                after=after,
                reason=f"PIM batch import run {self.run_id}",
            )
        except Exception:
            logger.exception("PimImporter audit failed sku=%s action=%s", sku, action)


def _safe_repr(value: Any) -> Any:
    """JSONB-safe representation — Decimal/datetime → str, dicts pasan tal cual."""
    from decimal import Decimal

    if isinstance(value, (Decimal, datetime)):
        return str(value)
    if isinstance(value, dict):
        return {k: _safe_repr(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_safe_repr(v) for v in value]
    return value


__all__ = ["COMMIT_EVERY_N_ROWS", "MAX_ERRORS_LOGGED", "PimImporter"]
