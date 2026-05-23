"""Parser openpyxl streaming para PIM completo.xlsx (US-1A-06-01).

Diseño:
- ``read_only=True`` + ``data_only=True`` → no carga las 5k filas en RAM.
- Sin custom_mapping: verifica header exacto contra :data:`EXPECTED_HEADERS`.
- Con custom_mapping: salta header_row_index filas, usa map_row_with_mapping.
- Detecta SKUs duplicados dentro del archivo (BR-1a-PIM-DUP).
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO

from app.services.importer.column_mapper import EXPECTED_HEADERS, map_row, map_row_with_mapping

if TYPE_CHECKING:
    from app.services.importer.mapping_detector import ColumnMappingItem


@dataclass(slots=True)
class ParsedRow:
    """Una fila parseada del PIM."""

    row_index: int  # 1-based, donde 1 = primera fila de datos (después del header)
    sku: str | None
    payload: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors and self.sku is not None


@dataclass(slots=True)
class ParseResult:
    """Resultado completo del parse."""

    rows: list[ParsedRow]
    header_errors: list[str]
    total_data_rows: int
    duplicate_skus: list[str]

    @property
    def header_ok(self) -> bool:
        return not self.header_errors


class HeaderMismatchError(ValueError):
    """El header del archivo no coincide con :data:`EXPECTED_HEADERS`."""


def _validate_header(header: tuple[Any, ...]) -> list[str]:
    """Valida que el header coincida exactamente con la spec sprint0."""
    errors: list[str] = []
    if len(header) < len(EXPECTED_HEADERS):
        errors.append(f"Archivo con {len(header)} columnas; esperadas {len(EXPECTED_HEADERS)}.")
        return errors
    for i, expected in enumerate(EXPECTED_HEADERS):
        actual = header[i]
        actual_str = (str(actual) if actual is not None else "").strip()
        if actual_str != expected:
            errors.append(f"col {i}: header esperado {expected!r}, recibido {actual_str!r}.")
    return errors


def parse_xlsx_stream(
    source: str | Path | BinaryIO,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
    header_row_index: int | None = None,
    custom_mapping: "list[ColumnMappingItem] | None" = None,
) -> ParseResult:
    """Parsea un xlsx PIM completo con openpyxl streaming.

    Args:
        source: path o file-like binario.
        sheet_name: nombre de sheet (default: la primera, alineado con el real).
        max_rows: límite de filas de datos a procesar (None = todas).
        header_row_index: si se provee, salta esas filas antes de leer
            el header. Útil cuando el xlsx tiene filas de título/metadatos.
        custom_mapping: lista de ColumnMappingItem. Si se provee, salta la
            validación de EXPECTED_HEADERS y usa map_row_with_mapping en
            lugar de map_row. header_row_index debe también proveerse si
            el xlsx tiene filas de título.
    """
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        sh = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        rows_iter: Iterator[tuple[Any, ...]] = sh.iter_rows(values_only=True)

        # Saltar filas de título/metadatos si se especifica.
        if header_row_index is not None:
            for _ in range(header_row_index):
                try:
                    next(rows_iter)
                except StopIteration:
                    return ParseResult(
                        rows=[],
                        header_errors=[
                            f"header_row_index={header_row_index} excede las filas del archivo."
                        ],
                        total_data_rows=0,
                        duplicate_skus=[],
                    )

        try:
            header = next(rows_iter)
        except StopIteration:
            return ParseResult(
                rows=[],
                header_errors=["Archivo vacío (sin header)."],
                total_data_rows=0,
                duplicate_skus=[],
            )

        # Validación de header (sólo si no hay custom_mapping).
        if custom_mapping is None:
            header_errors = _validate_header(header)
            if header_errors:
                return ParseResult(
                    rows=[], header_errors=header_errors, total_data_rows=0, duplicate_skus=[]
                )

        headers_list = [str(v).strip() if v is not None else "" for v in header]

        rows: list[ParsedRow] = []
        seen: dict[str, int] = {}
        duplicates: list[str] = []
        for i, row in enumerate(rows_iter, start=1):
            if max_rows is not None and i > max_rows:
                break
            # Skip filas totalmente vacías (openpyxl puede emitir tail rows None).
            if all(v is None or v == "" for v in row):
                continue

            if custom_mapping is not None:
                payload, errors = map_row_with_mapping(row, headers_list, custom_mapping)
            else:
                payload, errors = map_row(row, EXPECTED_HEADERS)

            sku = payload.get("sku")
            if sku is not None:
                if sku in seen:
                    duplicates.append(sku)
                    errors.append(f"SKU duplicado en archivo (primera ocurrencia row {seen[sku]}).")
                else:
                    seen[sku] = i
            rows.append(ParsedRow(row_index=i, sku=sku, payload=payload, errors=errors))

        return ParseResult(
            rows=rows,
            header_errors=[],
            total_data_rows=len(rows),
            duplicate_skus=duplicates,
        )
    finally:
        wb.close()
