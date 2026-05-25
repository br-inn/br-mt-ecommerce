"""Detección automática de estructura xlsx y propuesta de mapeo via LLM."""

from __future__ import annotations

import io
import json
import logging
import re
import unicodedata
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)

# Transforms disponibles (presentados en el frontend como opciones).
AVAILABLE_TRANSFORMS: tuple[str, ...] = (
    "text",
    "int",
    "decimal",
    "cm_to_mm",
    "ean",
    "bool_check",
    "percent",
)

# Campos target disponibles (escalares directos en products).
SCALAR_FIELDS: frozenset[str] = frozenset(
    {
        "sku",
        "family",
        "subfamily",
        "type",
        "erp_name",
        "intrastat_code",
        "hs_code",
        "connection",
        "brand",
        "weight",
        "bore_mm",
        "pressure_max_bar",
        "temp_min_c",
        "temp_max_c",
        "series",
        "material",
        "dn",
        "pn",
        "size",
        "revision",
        "external_url",
        "gtin",
        "dimensional_standard",
        "country_of_origin",
    }
)

# Prefijos JSONB válidos (el sufijo es la clave dentro del bucket).
JSONB_PREFIXES: frozenset[str] = frozenset({"dimensions", "packaging", "specs"})


@dataclass(frozen=True, slots=True)
class ColumnMappingItem:
    """Mapeo de una columna Excel a un campo de `products`."""

    excel_col: str
    target_field: str  # 'sku' | 'family' | 'dimensions.high_mm' | 'specs.ean_box' | '_skip'
    transform: str  # uno de AVAILABLE_TRANSFORMS
    confidence: float = 1.0
    notes: str = ""


def _is_header_row(row: tuple[Any, ...]) -> bool:
    """Heurística: una fila ES cabecera si tiene ≥3 celdas no-vacías y cortas.

    Descarta filas de título típicas ('PIM CONSOLIDADO...', 'Generado: ...').
    """
    non_empty = [v for v in row if v is not None and str(v).strip()]
    first = str(non_empty[0]).strip() if non_empty else ""
    if len(non_empty) < 3:
        # A row with < 3 non-empty cells is never a header.
        return False
    if first.upper().startswith(("PIM ", "GENERADO", "FUENTE")):
        return False
    # Si la primera celda es un número (fila de datos), no es cabecera.
    try:
        float(first)
        return False
    except ValueError:
        pass
    return True


def detect_header_row(
    xlsx_bytes: bytes,
    max_scan_rows: int = 10,
) -> tuple[int, list[str], list[list[Any]]]:
    """Detecta la fila de cabecera real en cualquier xlsx PIM.

    Itera las primeras `max_scan_rows` filas buscando la primera que tenga
    ≥3 celdas no-vacías y no parezca un título o metadato.

    Returns:
        (header_row_index, headers, sample_data_rows)
        - header_row_index: índice 0-based de la fila de cabecera.
        - headers: lista de nombres de columna (strings).
        - sample_data_rows: lista de hasta 5 filas de datos como listas.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        all_rows: list[tuple[Any, ...]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_scan_rows + 5:
                break
            all_rows.append(row)
    finally:
        wb.close()

    header_idx: int | None = None
    for i, row in enumerate(all_rows[:max_scan_rows]):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            f"No header row detected in the first {max_scan_rows} rows of the xlsx file."
        )

    headers_raw = all_rows[header_idx]
    headers = [str(v).strip() if v is not None else "" for v in headers_raw]
    # Strip trailing empty headers.
    while headers and not headers[-1]:
        headers.pop()

    # Collect up to 5 non-empty data rows after the header.
    samples: list[list[Any]] = []
    for row in all_rows[header_idx + 1 :]:
        if any(v is not None and v != "" for v in row):
            samples.append(list(row))
        if len(samples) >= 5:
            break

    return header_idx, headers, samples


def _normalize_header(s: str) -> str:
    """Lowercase + strip accents + collapse non-alphanumeric runs to single space."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


# Rules evaluated in order; first match wins.
# (compiled_pattern_on_normalized_header, target_field, transform, confidence)
_P = re.compile  # short alias for the list below
_HEURISTIC_RULES: list[tuple[re.Pattern[str], str, str, float]] = [
    # SKU / variant reference
    (
        _P(r"^sku$|^referencia\s+de\s+variante|^referencia\s+variante|^item\s+code"),
        "sku",
        "text",
        0.95,
    ),
    # Family hierarchy (subfamily before family to avoid early-exit)
    (_P(r"^subfamilia|^subfamily"), "subfamily", "text", 0.90),
    (_P(r"^familia$|^family$"), "family", "text", 0.90),
    # Trade classification codes
    (_P(r"^hs\s+code|^codigo\s+hs|^hs$"), "hs_code", "text", 0.85),
    (_P(r"^cod\s*intrastat|^intrastat\s+code|^codigo\s+intrastat"), "intrastat_code", "text", 0.85),
    # ERP name — must precede translations.en to avoid "nombre erp" matching "nombre en"
    (_P(r"^nombre\s+erp|^erp\s+name|^nombre\s+ax"), "erp_name", "text", 0.90),
    # Translations (only langs supported by the DB check constraint: en, es, ar)
    (_P(r"^nombre\s+es\b|^name\s+es\b|^descripcion\s+es\b"), "translations.es", "text", 0.95),
    (_P(r"^nombre\s+en\b|^name\s+en\b|^descripcion\s+en\b"), "translations.en", "text", 0.95),
    (_P(r"^nombre\s+ar\b|^name\s+ar\b"), "translations.ar", "text", 0.90),
    # EAN / GTIN — inner > box > individual > generic
    (_P(r"ean.*(inner|interna)|inner.*ean|ean\s+code\s+inner"), "specs.ean_inner_box", "ean", 0.95),
    (_P(r"ean.*(caja|box)|caja.*ean|ean\s+code\s+box"), "specs.ean_box", "ean", 0.90),
    (
        _P(r"ean.*(unidad|individual|unit)|individual.*ean|^individual\s+ean\s+code"),
        "specs.ean_individual",
        "ean",
        0.90,
    ),
    (_P(r"^ean\s+code$|^ean$"), "specs.ean_individual", "ean", 0.75),
    (_P(r"^gtin$"), "gtin", "ean", 0.90),
    # Weight — net first (canonical "net weight unit" and Spanish "peso neto")
    (_P(r"^net\s+weight|^peso\s+neto"), "weight", "decimal", 0.90),
    (_P(r"^weight\s+unit$"), "specs.weight_gross_kg", "decimal", 0.85),
    (_P(r"^weight$|^peso$"), "weight", "decimal", 0.75),
    # Piece dimensions with cm (Spanish file → cm_to_mm)
    (_P(r"^alto\s+pieza|^high\s+piece|^altura\s+pieza"), "dimensions.high_mm", "cm_to_mm", 0.90),
    (
        _P(r"^ancho\s+pieza|^wide\s+piece|^width\s+piece|^anchura\s+pieza"),
        "dimensions.wide_mm",
        "cm_to_mm",
        0.90,
    ),
    (
        _P(r"^largo\s+pieza|^deep\s+piece|^depth\s+piece|^longitud\s+pieza"),
        "dimensions.deep_mm",
        "cm_to_mm",
        0.90,
    ),
    # Piece dimensions already in mm (canonical file)
    (_P(r"^high\s+mm$"), "dimensions.high_mm", "decimal", 0.95),
    (_P(r"^wide\s+mm$"), "dimensions.wide_mm", "decimal", 0.95),
    (_P(r"^deep\s+mm$"), "dimensions.deep_mm", "decimal", 0.95),
    # Box dimensions (cm_to_mm)
    (_P(r"^alto\s+caja|^box\s+high|^altura\s+caja"), "packaging.box_high_mm", "cm_to_mm", 0.90),
    (
        _P(r"^ancho\s+caja|^box\s+wide|^box\s+width|^anchura\s+caja"),
        "packaging.box_wide_mm",
        "cm_to_mm",
        0.90,
    ),
    (
        _P(r"^largo\s+caja|^box\s+deep|^box\s+depth|^longitud\s+caja"),
        "packaging.box_deep_mm",
        "cm_to_mm",
        0.90,
    ),
    # Packaging counts
    (
        _P(r"^qty.{0,3}caja|^qty\s*x\s*box|^cantidad\s+caja|^uds\s+caja"),
        "packaging.qty_per_box",
        "int",
        0.90,
    ),
    (_P(r"^moq\s+inner|^inner\s+moq|^moq\s+inner\s+box"), "packaging.moq_inner_box", "int", 0.90),
    (_P(r"^x\s+pallet$|^pallet\s+qty$|^qty\s+pallet$"), "packaging.x_pallet", "int", 0.85),
    # Optional product attributes
    (_P(r"^marca$|^brand$"), "brand", "text", 0.85),
    (_P(r"^serie$|^series$"), "series", "text", 0.85),
    (_P(r"^material$"), "material", "text", 0.85),
    (_P(r"^connection$|^conexion$"), "connection", "text", 0.80),
    (_P(r"^tipo$|^type$"), "type", "text", 0.80),
    (_P(r"^dn$|^diametro\s+nominal|^nominal\s+diameter"), "dn", "text", 0.85),
    (_P(r"^pn$|^presion\s+nominal|^nominal\s+pressure"), "pn", "text", 0.85),
    (
        _P(r"^pais\s+origen|^country\s+of\s+origin|^country$|^origen$"),
        "country_of_origin",
        "text",
        0.80,
    ),
    (_P(r"^external\s+url|^url$"), "external_url", "text", 0.80),
    (_P(r"^revision$|^rev$"), "revision", "text", 0.75),
    (_P(r"^certif|^normas\b|^homolog"), "certifications", "text", 0.75),
]


def _suggest_mapping_heuristic(
    headers: list[str],
    sample_rows: list[list[Any]],
) -> list[ColumnMappingItem]:
    """Rule-based fallback mapping for common Spanish/English PIM column names."""
    results: list[ColumnMappingItem] = []
    for h in headers:
        if not h:
            results.append(
                ColumnMappingItem(
                    excel_col=h, target_field="_skip", transform="text", confidence=0.0
                )
            )
            continue
        norm = _normalize_header(h)
        matched_target: str | None = None
        matched_transform = "text"
        matched_confidence = 0.0
        for pattern, target, transform, confidence in _HEURISTIC_RULES:
            if pattern.search(norm):
                matched_target = target
                matched_transform = transform
                matched_confidence = confidence
                break
        if matched_target:
            results.append(
                ColumnMappingItem(
                    excel_col=h,
                    target_field=matched_target,
                    transform=matched_transform,
                    confidence=matched_confidence,
                    notes="Mapeo automático por reglas — verificar antes de importar",
                )
            )
        else:
            results.append(
                ColumnMappingItem(
                    excel_col=h,
                    target_field="_skip",
                    transform="text",
                    confidence=0.3,
                    notes="Sin mapeo automático — configurar manualmente",
                )
            )
    return results


_LLM_MODEL = "claude-sonnet-4-6"

_AVAILABLE_FIELDS_DOC = """
Scalar fields (products table):
  sku (required), family, subfamily, type, erp_name, intrastat_code, hs_code,
  connection, brand, weight, bore_mm, pressure_max_bar, temp_min_c, temp_max_c,
  series, material, dn, pn, size, revision, external_url, gtin,
  dimensional_standard, country_of_origin

JSONB sub-fields (dot notation):
  dimensions.high_mm, dimensions.wide_mm, dimensions.deep_mm
  packaging.qty_per_box, packaging.box_high_mm, packaging.box_wide_mm,
  packaging.box_deep_mm, packaging.moq_inner_box, packaging.x_pallet
  specs.<any_key>   ← use for EANs, booleans, flags not covered by scalars

Translations (writes to product_translations.name for that lang):
  translations.en   translations.es   translations.fr
  translations.de   translations.it   translations.pt   translations.ar

Multi-value comma-separated (M:N):
  certifications    ← "CE, ISO 9001, WRAS" is split + each resolved to vocab

Special:
  _skip             ← ignore this column entirely

Available transforms:
  text        plain text / string
  int         integer number
  decimal     decimal / float
  cm_to_mm    multiply x10 (centimeters to millimeters)
  ean         EAN barcode (digits only, valid lengths 8/12/13/14)
  bool_check  truthy check: "✓", "yes", "1", "true" → true; else false
  percent     numeric percentage stored as integer 0-100
"""


async def suggest_mapping(
    headers: list[str],
    sample_rows: list[list[Any]],
    api_key: str = "",
) -> list[ColumnMappingItem]:
    """Llama a Claude para proponer el mapeo de columnas Excel → campos product.

    Devuelve lista de _skip (confidence=0) si el LLM falla o la key no está
    disponible, permitiendo al usuario mapear manualmente.
    """
    import anthropic

    if not api_key:
        logger.info("suggest_mapping: ANTHROPIC_API_KEY no configurada — usando mapeo heurístico")
        return _suggest_mapping_heuristic(headers, sample_rows)

    # Tabla columna → valores de muestra (todas las columnas, no truncadas).
    col_samples: list[str] = []
    for j, h in enumerate(headers):
        vals = [repr(row[j]) if j < len(row) else "None" for row in sample_rows[:3]]
        col_samples.append(f"  {h!r}: {', '.join(vals)}")
    samples_text = "\n".join(col_samples)

    prompt = (
        f"You are a product data mapping assistant for an industrial PVF "
        f"(pipes, valves, fittings) manufacturer PIM system.\n\n"
        f"Given these Excel column headers and their sample values, propose the "
        f"best mapping from each Excel column to a product database field.\n\n"
        f"Use the actual sample values to infer the correct field and transform. "
        f"For example: a column with values like 'DN25', 'DN40' maps to 'dn' with "
        f"transform 'text'; a column with values in cm (e.g. 12.5) that represents "
        f"a physical dimension maps to dimensions.* with transform 'cm_to_mm'.\n"
        f"For multi-language name columns (e.g. 'Nombre ES', 'Name EN', 'Nome IT'), "
        f"use translations.<lang> (translations.es, translations.en, translations.it, etc.).\n"
        f"For certification columns (e.g. 'Normas', 'Certifications', 'CE Mark', "
        f"'Homologaciones'), use 'certifications'.\n\n"
        f"{_AVAILABLE_FIELDS_DOC}\n\n"
        f"Column headers with sample values (up to 3 rows):\n{samples_text}\n\n"
        f"Return a JSON array — no markdown, no explanation, just JSON. "
        f"Each element:\n"
        f'  {{"excel_col": "<exact header>", "target_field": "<field>", '
        f'"transform": "<transform>", "confidence": 0.0-1.0, '
        f'"notes": "<1-sentence Spanish explanation>"}}\n\n'
        f"Include ALL {len(headers)} columns, even those mapped to _skip."
    )

    try:
        client = anthropic.AsyncAnthropic(api_key=api_key)
        message = await client.messages.create(
            model=_LLM_MODEL,
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}],
        )
        text = message.content[0].text.strip()  # type: ignore[union-attr]
        # Strip markdown code fences (e.g. ```json ... ```)
        text = re.sub(r"^```[^\n]*\n?", "", text, flags=re.MULTILINE).strip()
        data = json.loads(text)
        if not isinstance(data, list):
            logger.warning(
                "suggest_mapping: LLM returned %s instead of a JSON array"
                " — returning empty mapping",
                type(data).__name__,
            )
            return []
        return [
            ColumnMappingItem(
                excel_col=item["excel_col"],
                target_field=item.get("target_field", "_skip"),
                transform=item.get("transform", "text"),
                confidence=float(item.get("confidence", 0.5)),
                notes=item.get("notes", ""),
            )
            for item in data
            if isinstance(item, dict) and "excel_col" in item
        ]
    except Exception:
        logger.exception("suggest_mapping LLM call failed — returning _skip fallback")
        return [
            ColumnMappingItem(
                excel_col=h,
                target_field="_skip",
                transform="text",
                confidence=0.0,
                notes="LLM no disponible — mapeo manual requerido",
            )
            for h in headers
        ]
