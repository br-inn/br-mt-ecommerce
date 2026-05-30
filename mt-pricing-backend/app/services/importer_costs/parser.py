"""Parser openpyxl streaming para Excel de costos batch (US-1A-06-02).

Estructura mínima esperada del Excel (header literal en row 1):
    sku | scheme_code | supplier_code | currency | total
    | breakdown_fob | breakdown_freight | breakdown_customs
    | breakdown_fba_fee | breakdown_fbm_fee | breakdown_payment_fee
    | breakdown_marketing | breakdown_storage | breakdown_ppc
    | breakdown_otros | valid_from

Las columnas ``breakdown_*`` son opcionales (vacío = 0). ``valid_from`` es la
fecha de inicio de vigencia del coste (ISO ``YYYY-MM-DD`` o fecha Excel) y sirve
como ancla as-of del FX. Si no viene, se usa la fecha de hoy (``date.today()``).
Filas con ``valid_from`` futuro son válidas: crean un rango futuro y el
auto-encadenado del service cierra el rango abierto previo en ``valid_from - 1``.

Notas:
- ``read_only=True`` + ``data_only=True`` → no carga el archivo entero en RAM.
- Detecta SKUs+scheme+supplier duplicados (el applier los reportaría igual, pero
  preferimos cortar en parse).
- Cast estricto: ``total`` y cada ``breakdown_*`` → Decimal; vacío → None.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO

EXPECTED_COSTS_HEADERS: tuple[str, ...] = (
    "sku",
    "scheme_code",
    "supplier_code",
    "currency",
    "total",
    "breakdown_fob",
    "breakdown_freight",
    "breakdown_customs",
    "breakdown_fba_fee",
    "breakdown_fbm_fee",
    "breakdown_payment_fee",
    "breakdown_marketing",
    "breakdown_storage",
    "breakdown_ppc",
    "breakdown_otros",
    "valid_from",
)

BREAKDOWN_COLUMNS: tuple[str, ...] = tuple(
    h for h in EXPECTED_COSTS_HEADERS if h.startswith("breakdown_")
)


@dataclass(slots=True)
class CostRow:
    """Una fila parseada del Excel de costos."""

    row_index: int  # 1-based, 1 = primera fila de datos
    sku: str | None
    scheme_code: str | None
    supplier_code: str | None
    currency: str | None
    total: Decimal | None
    breakdown: dict[str, Any] = field(default_factory=dict)
    valid_from: date | None = None
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors and self.sku is not None and self.scheme_code is not None

    @property
    def dedup_key(self) -> tuple[str, str, str]:
        return (self.sku or "", self.scheme_code or "", self.supplier_code or "")


@dataclass(slots=True)
class CostsParseResult:
    rows: list[CostRow]
    header_errors: list[str]
    total_data_rows: int
    duplicate_keys: list[tuple[str, str, str]]

    @property
    def header_ok(self) -> bool:
        return not self.header_errors


def _validate_header(header: tuple[Any, ...]) -> list[str]:
    errors: list[str] = []
    if len(header) < len(EXPECTED_COSTS_HEADERS):
        errors.append(
            f"Archivo con {len(header)} columnas; esperadas {len(EXPECTED_COSTS_HEADERS)}."
        )
        return errors
    for i, expected in enumerate(EXPECTED_COSTS_HEADERS):
        actual = header[i]
        actual_str = (str(actual) if actual is not None else "").strip()
        if actual_str != expected:
            errors.append(f"col {i}: header esperado {expected!r}, recibido {actual_str!r}.")
    return errors


def _cast_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _cast_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Decimal inválido: {v!r}") from exc


def _cast_date(v: Any) -> date | None:
    """Parsea ``valid_from`` a ``date``. None/"" → None (el caller usa hoy).

    Acepta fechas Excel nativas (``datetime``/``date``) e ISO ``YYYY-MM-DD``
    (también con componente de hora, que se descarta).
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha inválida: {v!r}")


def _map_cost_row(excel_row: tuple[Any, ...] | list[Any], row_index: int) -> CostRow:
    errors: list[str] = []
    breakdown: dict[str, Any] = {}

    # Normaliza acceso por posición.
    def _at(idx: int) -> Any:
        return excel_row[idx] if idx < len(excel_row) else None

    # Required-ish: sku, scheme_code.
    sku = _cast_text(_at(0))
    scheme = _cast_text(_at(1))
    supplier = _cast_text(_at(2))
    currency = _cast_text(_at(3)) or "AED"
    if currency and len(currency) != 3:
        errors.append(f"currency inválida ({currency!r}); debe ser 3 letras.")
        currency = None

    try:
        total = _cast_decimal(_at(4))
    except ValueError as exc:
        errors.append(f"col 'total': {exc}")
        total = None

    if total is None:
        errors.append("col 'total': requerido y vino vacío.")
    elif total < 0:
        errors.append(f"col 'total': debe ser >= 0; vino {total}.")

    # Breakdown columns (idx 5..14).
    for i, key in enumerate(BREAKDOWN_COLUMNS, start=5):
        try:
            val = _cast_decimal(_at(i))
        except ValueError as exc:
            errors.append(f"col {key!r}: {exc}")
            continue
        if val is not None:
            short_key = key.replace("breakdown_", "")
            breakdown[short_key] = str(val)

    # valid_from idx 15. Si la fila no la trae, default a hoy (documentado).
    try:
        valid_from = _cast_date(_at(15))
    except ValueError as exc:
        errors.append(f"col 'valid_from': {exc}")
        valid_from = None
    else:
        if valid_from is None:
            valid_from = date.today()

    if sku is None:
        errors.append("col 'sku': requerido y vino vacío.")
    if scheme is None:
        errors.append("col 'scheme_code': requerido y vino vacío.")

    return CostRow(
        row_index=row_index,
        sku=sku,
        scheme_code=scheme,
        supplier_code=supplier,
        currency=currency,
        total=total,
        breakdown=breakdown,
        valid_from=valid_from,
        errors=errors,
    )


def parse_costs_xlsx_stream(
    source: str | Path | BinaryIO,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> CostsParseResult:
    """Parsea un xlsx de costos batch."""
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        sh = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows_iter: Iterator[tuple[Any, ...]] = sh.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return CostsParseResult(
                rows=[],
                header_errors=["Archivo vacío (sin header)."],
                total_data_rows=0,
                duplicate_keys=[],
            )

        header_errors = _validate_header(header)
        if header_errors:
            return CostsParseResult(
                rows=[],
                header_errors=header_errors,
                total_data_rows=0,
                duplicate_keys=[],
            )

        rows: list[CostRow] = []
        seen: dict[tuple[str, str, str], int] = {}
        duplicates: list[tuple[str, str, str]] = []
        for i, row in enumerate(rows_iter, start=1):
            if max_rows is not None and i > max_rows:
                break
            if all(v is None or v == "" for v in row):
                continue
            cr = _map_cost_row(row, i)
            key = cr.dedup_key
            if cr.sku is not None and cr.scheme_code is not None:
                if key in seen:
                    duplicates.append(key)
                    cr.errors.append(f"Duplicado en archivo (primera ocurrencia row {seen[key]}).")
                else:
                    seen[key] = i
            rows.append(cr)

        return CostsParseResult(
            rows=rows,
            header_errors=[],
            total_data_rows=len(rows),
            duplicate_keys=duplicates,
        )
    finally:
        wb.close()
