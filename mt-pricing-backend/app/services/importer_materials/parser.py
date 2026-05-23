"""Parser openpyxl streaming para Excel de compatibilidades materiales (US-1A-06-03).

Estructura mínima esperada (header row 1):

    producto_descriptor | temperatura_c | <material_1> | <material_2> | ...

- ``producto_descriptor`` TEXT requerido.
- ``temperatura_c`` NUMERIC requerido.
- Cualquier columna posterior se considera un "material" cuyo header normaliza
  a snake_case (``Acero Inoxidable 316L`` → ``acero_inoxidable_316l``). El
  valor de la celda se persiste como string normalizado:
    "OK" / "X" / "-" / "" → "ok" / "x" / "-" / None.

NO valida el set de materiales contra una whitelist — la tabla
``material_compatibilities`` es referencial, no maestra.
"""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Iterator
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO


@dataclass(slots=True)
class MaterialRow:
    row_index: int
    producto_descriptor: str | None
    temperatura_c: Decimal | None
    compatibilities: dict[str, str] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return (
            not self.errors
            and self.producto_descriptor is not None
            and self.temperatura_c is not None
        )


@dataclass(slots=True)
class MaterialsParseResult:
    rows: list[MaterialRow]
    header_errors: list[str]
    materials_columns: list[str]  # snake-cased column names (order of Excel)
    total_data_rows: int

    @property
    def header_ok(self) -> bool:
        return not self.header_errors


_REQUIRED_LEFT_HEADERS: tuple[str, ...] = ("producto_descriptor", "temperatura_c")


def _normalize_header(s: str) -> str:
    """Normaliza header del Excel a snake_case ASCII."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # Reemplaza caracteres acentuados por su equivalente ASCII.
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    # Substituye separadores y caracteres no alfanuméricos por _ .
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s


def _validate_left_headers(header: tuple[Any, ...]) -> tuple[list[str], list[str]]:
    """Devuelve (errors, normalized_full_header)."""
    errors: list[str] = []
    if len(header) < 3:
        errors.append(
            f"Archivo con {len(header)} columnas; al menos 3 (descriptor, temp, "
            f"un material) son requeridas."
        )
        return errors, []
    norm = [_normalize_header(h) for h in header]
    for i, expected in enumerate(_REQUIRED_LEFT_HEADERS):
        if i >= len(norm) or norm[i] != expected:
            errors.append(f"col {i}: header esperado {expected!r}, recibido {norm[i]!r}.")
    return errors, norm


def _cast_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Decimal inválido: {v!r}") from exc


def _cast_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _normalize_compat(v: Any) -> str | None:
    """Normaliza el flag de compatibilidad — None si vacío."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    return s


def parse_materials_xlsx_stream(
    source: str | Path | BinaryIO,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> MaterialsParseResult:
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        sh = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows_iter: Iterator[tuple[Any, ...]] = sh.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return MaterialsParseResult(
                rows=[],
                header_errors=["Archivo vacío (sin header)."],
                materials_columns=[],
                total_data_rows=0,
            )
        header_errors, normalized = _validate_left_headers(header)
        if header_errors:
            return MaterialsParseResult(
                rows=[],
                header_errors=header_errors,
                materials_columns=[],
                total_data_rows=0,
            )
        materials_cols: list[str] = []
        for col_idx in range(2, len(normalized)):
            col = normalized[col_idx]
            if not col:
                continue
            materials_cols.append(col)

        rows: list[MaterialRow] = []
        for i, row in enumerate(rows_iter, start=1):
            if max_rows is not None and i > max_rows:
                break
            if all(v is None or v == "" for v in row):
                continue

            errors: list[str] = []
            descriptor = _cast_text(row[0] if len(row) > 0 else None)
            try:
                temp = _cast_decimal(row[1] if len(row) > 1 else None)
            except ValueError as exc:
                errors.append(f"col 'temperatura_c': {exc}")
                temp = None

            if descriptor is None:
                errors.append("col 'producto_descriptor': requerido y vino vacío.")
            if temp is None and "col 'temperatura_c'" not in "; ".join(errors):
                errors.append("col 'temperatura_c': requerido y vino vacío.")

            compats: dict[str, str] = {}
            for j, mat in enumerate(materials_cols, start=2):
                val = row[j] if j < len(row) else None
                norm = _normalize_compat(val)
                if norm is not None:
                    compats[mat] = norm

            rows.append(
                MaterialRow(
                    row_index=i,
                    producto_descriptor=descriptor,
                    temperatura_c=temp,
                    compatibilities=compats,
                    errors=errors,
                )
            )

        return MaterialsParseResult(
            rows=rows,
            header_errors=[],
            materials_columns=materials_cols,
            total_data_rows=len(rows),
        )
    finally:
        wb.close()
