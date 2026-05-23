"""Seed 50 demo products covering Stage 3 taxonomy diversity (Wave 11).

Reads `catalogo_mt_productos.xlsx` (4181 rows: Sección · Material · Categoría ·
Código · Medida · Página) and `PIM completo.xlsx` (5086 rows: variant ref,
EAN, weight, dimensions, packaging). Picks 50 representative codes covering:

- 7 Secciones (Válvulas y Filtros · Accesorios y Bridas · Automatización ·
  Recambios · Instrumentos Medición · Accesorios Sanitarios · Otros).
- 9 Materiales (Latón · Acero Inoxidable · Galvanizado · Fundición · PPR ·
  Plástico/PVC · Equipos · Otros · Varios).
- 53 Categorías (top 25 covered).

Para cada producto:
- Crea/asegura la jerarquía Family → Subfamily → ProductType (taxonomía Opción C).
- Crea/asegura la Series si la Categoría sugiere una serie comercial (MT Press,
  PN40 Platinum, etc.).
- Resuelve material_id contra el vocabulario curado (laton / acero_inoxidable / …).
- Asigna divisiones M:N: Hidrosanitario por defecto; Industrial si Material ∈
  {Acero Inoxidable, Fundición, Equipos}.
- Hace UPSERT del producto con todos los campos Stage 3 + traducción ES.

Idempotente: re-run actualiza los 50 sin duplicar.

Uso:
    docker cp Documentos\\ referencia\\ de\\ articulos/catalogo_mt_productos.xlsx mt-backend:/tmp/catalogo.xlsx
    docker cp Documentos\\ referencia\\ de\\ articulos/PIM\\ completo.xlsx mt-backend:/tmp/pim.xlsx
    docker exec mt-backend python -m scripts.seed_demo_50_products
"""

from __future__ import annotations

import asyncio
import re
import unicodedata
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.engine import get_sessionmaker
from app.db.models.product import Product, ProductTranslation
from app.db.models.vocabularies import (
    Brand,
    Division,
    Family,
    Material,
    ProductDivision,
    ProductType,
    Series,
    SeriesTier,
    Subfamily,
)

CATALOG_PATH = Path("/tmp/catalogo.xlsx")
PIM_PATH = Path("/tmp/pim.xlsx")

# ---------------------------------------------------------------------------
# Mapping tables (Sección/Material/Categoría → taxonomy codes)
# ---------------------------------------------------------------------------

SECTION_TO_FAMILY: dict[str, tuple[str, str]] = {
    # spanish_seccion → (family_code, family_name_en)
    "Válvulas y Filtros": ("valves_and_filters", "Valves and Filters"),
    "Accesorios y Bridas": ("fittings_and_flanges", "Fittings and Flanges"),
    "Automatización": ("valves_automation", "Valves Automation"),
    "Recambios": ("spare_parts", "Spare Parts"),
    "Instrumentos Medición": ("measuring_instruments", "Measuring Instruments"),
    "Accesorios Sanitarios": ("sanitary_accessories", "Sanitary Accessories"),
    "Otros": ("other", "Other"),
}

MATERIAL_ES_TO_CODE: dict[str, str | None] = {
    "Latón": "laton",
    "Acero Inoxidable": "acero_inoxidable",
    "Galvanizado": "galvanizado",
    "Fundición": "fundicion",
    "PPR": "ppr",
    "Plástico / PVC": "plastico_pvc",
    "Equipos": None,  # actuator equipment, no curated material
    "Otros": None,
    "Varios": None,
}

# Material → divisions M:N. Default Hidrosanitario; heavy-duty add Industrial.
MATERIAL_TO_DIVISIONS: dict[str, list[str]] = {
    "Latón": ["hidrosanitario"],
    "Acero Inoxidable": ["hidrosanitario", "industrial"],
    "Galvanizado": ["hidrosanitario"],
    "Fundición": ["industrial"],
    "PPR": ["hidrosanitario"],
    "Plástico / PVC": ["hidrosanitario"],
    "Equipos": ["industrial"],
    "Otros": ["hidrosanitario"],
    "Varios": ["hidrosanitario"],
}

# Series detection from Categoría. Returns (code, name_en, tier_code, pressure_pn).
SERIES_FROM_CATEGORY: dict[str, tuple[str, str, str | None, int | None]] = {
    "MT PRESS SYSTEM": ("mt_press_system", "MT Press System", "platinum", 16),
    "MULTILAYER PIPELINE": ("multilayer_pipeline", "Multilayer Pipeline", None, None),
    "MULTILAYER BALL VALVES": ("multilayer_pipeline", "Multilayer Pipeline", None, None),
    "PERT-AL-PERT": ("pert_al_pert", "PERT-AL-PERT", None, None),
    "GARDEN TAPS": ("mt_garden", "MT Garden", "silver", 10),
    "DVGW | WATERMARK": ("pn40_platinum", "PN40 Platinum Series", "platinum", 40),
    "PZH | WRAS | ACS": ("pn30_gold", "PN30 Gold Series", "gold", 30),
    "THREADED BALL VALVES": ("threaded_ball", "Threaded Ball Valves", None, None),
    "BRASS BALL VALVES": ("threaded_ball", "Threaded Ball Valves", None, None),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def slugify(s: str) -> str:
    """Lowercase + strip accents + replace non-alnum with underscores."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s or "other"


def medida_to_dn(medida: str | None) -> str | None:
    """Convert catalog Medida like '1/2"' or '50' to canonical DN like '15' or '50'.

    Mapeo aproximado del DN según ISO 6708 — equivalencias estandar pulgada→DN.
    """
    if not medida:
        return None
    s = str(medida).replace('"', "").replace("”", "").strip()
    # Direct numeric DN (already in mm)
    try:
        n = float(s)
        if n >= 8 and n <= 1000:
            # Whole DN values
            for dn in [8, 10, 15, 20, 25, 32, 40, 50, 65, 80, 100, 125, 150, 200, 250, 300]:
                if abs(n - dn) < 0.5:
                    return str(dn)
    except ValueError:
        pass
    # Imperial inch fractions → DN map
    inch_to_dn = {
        "1/4": "8",
        "3/8": "10",
        "1/2": "15",
        "3/4": "20",
        "1": "25",
        "1 1/4": "32",
        "1.1/4": "32",
        "1 1/2": "40",
        "1.1/2": "40",
        "2": "50",
        "2 1/2": "65",
        "2.1/2": "65",
        "3": "80",
        "4": "100",
        "5": "125",
        "6": "150",
        "8": "200",
    }
    if s in inch_to_dn:
        return inch_to_dn[s]
    return None


def medida_to_size(medida: str | None) -> str | None:
    """Devuelve la pulgada original como `size` legible: '1/2"', '3"', etc."""
    if not medida:
        return None
    s = str(medida).strip()
    if not s.endswith('"') and not s.endswith("”"):
        s = s + '"'
    return s


# ---------------------------------------------------------------------------
# Xlsx readers
# ---------------------------------------------------------------------------


def load_catalog() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(str(CATALOG_PATH), read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        sec, mat, cat, cod, med, pag = r
        if not cod or not str(cod).strip():
            continue
        rows.append(
            {
                "seccion": sec,
                "material": mat,
                "categoria": cat,
                "codigo": str(cod).strip(),
                "medida": str(med) if med else None,
                "pagina": pag,
            }
        )
    return rows


def load_pim_index() -> dict[str, dict[str, Any]]:
    """Index PIM by `Referencia de variante` (which equals catalog Código)."""
    wb = openpyxl.load_workbook(str(PIM_PATH), read_only=True, data_only=True)
    ws = wb["sheet1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx: dict[str, dict[str, Any]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        ref = str(r[0]).strip()
        idx[ref] = dict(zip(headers, r, strict=False))
    return idx


# ---------------------------------------------------------------------------
# Selection: pick 50 with diversity
# ---------------------------------------------------------------------------


def medida_sort_key(med: str | None) -> float:
    if not med:
        return 99999
    s = str(med).replace('"', "").replace("”", "").strip()
    try:
        if "/" in s:
            num, den = s.split("/", 1)
            return float(num) / float(den)
        return float(s)
    except ValueError:
        return 99999


def pick_50(catalog_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_combo: dict[tuple[Any, Any, Any], list[dict[str, Any]]] = defaultdict(list)
    for r in catalog_rows:
        by_combo[(r["seccion"], r["material"], r["categoria"])].append(r)

    selected: list[dict[str, Any]] = []
    for combo, lst in by_combo.items():
        lst_sorted = sorted(lst, key=lambda r: medida_sort_key(r["medida"]))
        # Pick smallest size from each combo
        selected.append(lst_sorted[0])

    # Sort to spread sections evenly
    selected.sort(key=lambda r: (r["seccion"], r["material"], r["categoria"]))

    # If too many, downsample uniformly to 50
    if len(selected) > 50:
        step = len(selected) / 50.0
        picked = [selected[int(i * step)] for i in range(50)]
        # Dedupe (round may collide)
        seen: set[str] = set()
        out: list[dict[str, Any]] = []
        for r in picked:
            if r["codigo"] not in seen:
                seen.add(r["codigo"])
                out.append(r)
        # Top-up if dedupe shrunk under 50
        i = 0
        while len(out) < 50 and i < len(selected):
            if selected[i]["codigo"] not in seen:
                seen.add(selected[i]["codigo"])
                out.append(selected[i])
            i += 1
        return out[:50]
    return selected


# ---------------------------------------------------------------------------
# DB upsert helpers
# ---------------------------------------------------------------------------


async def ensure_division(s: AsyncSession, code: str, name: str) -> Division:
    res = await s.execute(select(Division).where(Division.code == code))
    row = res.scalar_one_or_none()
    if row:
        return row
    row = Division(code=code, name=name)
    s.add(row)
    await s.flush()
    return row


async def ensure_brand(s: AsyncSession, code: str, name: str) -> Brand:
    res = await s.execute(select(Brand).where(Brand.code == code))
    row = res.scalar_one_or_none()
    if row:
        return row
    row = Brand(code=code, name=name)
    s.add(row)
    await s.flush()
    return row


async def ensure_family(s: AsyncSession, code: str, name: str) -> Family:
    res = await s.execute(select(Family).where(Family.code == code))
    row = res.scalar_one_or_none()
    if row:
        return row
    row = Family(code=code, name=name)
    s.add(row)
    await s.flush()
    return row


async def ensure_subfamily(s: AsyncSession, family_id: UUID, code: str, name: str) -> Subfamily:
    res = await s.execute(
        select(Subfamily).where(Subfamily.family_id == family_id, Subfamily.code == code)
    )
    row = res.scalar_one_or_none()
    if row:
        return row
    row = Subfamily(family_id=family_id, code=code, name=name)
    s.add(row)
    await s.flush()
    return row


async def ensure_product_type(
    s: AsyncSession, subfamily_id: UUID, code: str, name: str
) -> ProductType:
    res = await s.execute(
        select(ProductType).where(
            ProductType.subfamily_id == subfamily_id, ProductType.code == code
        )
    )
    row = res.scalar_one_or_none()
    if row:
        return row
    row = ProductType(subfamily_id=subfamily_id, code=code, name=name)
    s.add(row)
    await s.flush()
    return row


async def ensure_series(
    s: AsyncSession,
    code: str,
    name_en: str,
    tier_code: str | None,
    pressure_pn: int | None,
) -> Series:
    res = await s.execute(select(Series).where(Series.code == code))
    row = res.scalar_one_or_none()
    if row:
        return row
    tier_id: UUID | None = None
    if tier_code:
        tres = await s.execute(select(SeriesTier).where(SeriesTier.code == tier_code))
        tier = tres.scalar_one_or_none()
        if tier:
            tier_id = tier.id
    row = Series(
        code=code,
        name_en=name_en,
        tier_id=tier_id,
        pressure_rating_pn=pressure_pn,
    )
    s.add(row)
    await s.flush()
    return row


async def find_material_id(s: AsyncSession, code: str | None) -> UUID | None:
    if not code:
        return None
    res = await s.execute(
        select(Material.id).where(Material.code == code, Material.active.is_(True))
    )
    return res.scalar_one_or_none()


async def upsert_product(
    s: AsyncSession,
    *,
    sku: str,
    name_en: str,
    family: str,
    subfamily: str | None,
    type_: str | None,
    material_text: str | None,
    dn: str | None,
    pn: str | None,
    series_text: str | None,
    family_id: UUID,
    subfamily_id: UUID | None,
    type_id: UUID | None,
    brand_id: UUID,
    series_id: UUID | None,
    material_id: UUID | None,
    weight: Decimal | None,
    weight_unit: str | None,
    intrastat_code: str | None,
    erp_name: str | None,
    description_en: str | None,
    packaging: dict[str, Any],
    dimensions: dict[str, Any],
    dn_text: str | None,
) -> Product:
    res = await s.execute(select(Product).where(Product.sku == sku))
    p = res.scalar_one_or_none()
    if p:
        # Update Stage 3 fields (idempotent re-run)
        p.name_en = name_en
        p.family = family
        p.subfamily = subfamily
        p.type = type_
        p.material = material_text
        p.dn = dn
        p.pn = pn
        p.series = series_text
        p.family_id = family_id
        p.subfamily_id = subfamily_id
        p.type_id = type_id
        p.brand_id = brand_id
        p.series_id = series_id
        p.material_id = material_id
        if weight is not None:
            p.weight = weight
        if weight_unit:
            p.weight_unit = weight_unit
        if intrastat_code:
            p.intrastat_code = intrastat_code
        if erp_name:
            p.erp_name = erp_name
        if description_en:
            p.description_en = description_en
        if packaging:
            p.packaging = packaging
        if dimensions:
            p.dimensions = dimensions
        await s.flush()
        return p
    p = Product(
        sku=sku,
        name_en=name_en,
        family=family,
        subfamily=subfamily,
        type=type_,
        material=material_text,
        dn=dn,
        pn=pn,
        series=series_text,
        family_id=family_id,
        subfamily_id=subfamily_id,
        type_id=type_id,
        brand_id=brand_id,
        series_id=series_id,
        material_id=material_id,
        weight=weight,
        weight_unit=weight_unit,
        intrastat_code=intrastat_code,
        erp_name=erp_name,
        description_en=description_en,
        packaging=packaging,
        dimensions=dimensions,
        active=True,
        data_quality="complete",
    )
    s.add(p)
    await s.flush()
    return p


async def ensure_translation_es(
    s: AsyncSession,
    sku: str,
    name: str,
    description: str | None,
) -> None:
    res = await s.execute(
        select(ProductTranslation).where(
            ProductTranslation.sku == sku, ProductTranslation.lang == "es"
        )
    )
    tr = res.scalar_one_or_none()
    if tr:
        tr.name = name
        if description:
            tr.description = description
        return
    tr = ProductTranslation(
        sku=sku,
        lang="es",
        name=name,
        description=description,
        status="approved",
    )
    s.add(tr)


async def ensure_product_division(s: AsyncSession, sku: str, division_id: UUID) -> None:
    res = await s.execute(
        select(ProductDivision).where(
            ProductDivision.product_sku == sku,
            ProductDivision.division_id == division_id,
        )
    )
    if res.scalar_one_or_none():
        return
    s.add(ProductDivision(product_sku=sku, division_id=division_id))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


async def main() -> None:
    catalog = load_catalog()
    pim = load_pim_index()
    print(f"Catalog rows: {len(catalog)} | PIM index: {len(pim)}")
    selected = pick_50(catalog)
    print(f"Selected: {len(selected)} products covering Stage 3 diversity")

    sm = get_sessionmaker()
    async with sm() as s:
        # Pre-load divisions
        div_id_by_code: dict[str, UUID] = {}
        for code, name in [("hidrosanitario", "Hidrosanitario"), ("industrial", "Industrial")]:
            div = await ensure_division(s, code, name)
            div_id_by_code[code] = div.id

        brand = await ensure_brand(s, "mt", "MT")

        ok = 0
        skipped: list[str] = []
        for r in selected:
            sku = r["codigo"]
            seccion = r["seccion"] or "Otros"
            material_es = r["material"] or "Otros"
            categoria = r["categoria"] or "OTHER"

            family_code, family_name = SECTION_TO_FAMILY.get(seccion, ("other", "Other"))
            family = await ensure_family(s, family_code, family_name)

            sub_code = slugify(material_es)
            subfamily = await ensure_subfamily(s, family.id, sub_code, material_es)
            type_code = slugify(categoria)[:64]
            ptype = await ensure_product_type(s, subfamily.id, type_code, categoria.title())

            mat_code = MATERIAL_ES_TO_CODE.get(material_es)
            mat_id = await find_material_id(s, mat_code) if mat_code else None

            series_id: UUID | None = None
            series_text: str | None = None
            for cat_pattern, (scode, sname, tcode, ppn) in SERIES_FROM_CATEGORY.items():
                if cat_pattern in (categoria or "").upper():
                    ser = await ensure_series(s, scode, sname, tcode, ppn)
                    series_id = ser.id
                    series_text = scode
                    break

            pim_row = pim.get(sku, {})
            pim_name = pim_row.get("Nombre ERP - AX")
            name_en = (pim_name or f"{categoria} {material_es} {r['medida'] or ''}").strip()
            intrastat = pim_row.get("Cod.Intrastat - AX")
            weight_raw = pim_row.get("net weight unit") or pim_row.get("weight unit")
            try:
                weight = Decimal(str(weight_raw)) if weight_raw else None
            except Exception:
                weight = None
            packaging: dict[str, Any] = {}
            if pim_row.get("qty x box"):
                packaging["qty_per_box"] = pim_row["qty x box"]
            if pim_row.get("MOQ INNER BOX"):
                packaging["qty_per_inner_box"] = pim_row["MOQ INNER BOX"]
            if pim_row.get("X PALLET"):
                packaging["qty_per_pallet"] = pim_row["X PALLET"]
            if pim_row.get("INDIVIDUAL EAN CODE"):
                packaging["ean_individual"] = str(pim_row["INDIVIDUAL EAN CODE"])
            dimensions: dict[str, Any] = {}
            for src, key in [
                ("High mm", "height_mm"),
                ("Wide mm", "width_mm"),
                ("Deep mm", "depth_mm"),
            ]:
                if pim_row.get(src):
                    dimensions[key] = pim_row[src]

            dn = medida_to_dn(r["medida"])
            size_text = medida_to_size(r["medida"])

            await upsert_product(
                s,
                sku=sku,
                name_en=name_en,
                family=family_code,
                subfamily=sub_code,
                type_=type_code,
                material_text=mat_code or material_es.lower().replace("/", "_").replace(" ", "_"),
                dn=dn,
                pn=None,
                series_text=series_text,
                family_id=family.id,
                subfamily_id=subfamily.id,
                type_id=ptype.id,
                brand_id=brand.id,
                series_id=series_id,
                material_id=mat_id,
                weight=weight,
                weight_unit="kg",
                intrastat_code=str(intrastat) if intrastat else None,
                erp_name=pim_name,
                description_en=f"{categoria.title()} — {material_es} · medida {size_text or '?'}",
                packaging=packaging,
                dimensions=dimensions,
                dn_text=size_text,
            )
            await ensure_translation_es(
                s,
                sku=sku,
                name=f"{categoria.title()} {material_es} {size_text or ''}".strip(),
                description=f"Sección {seccion} · {material_es} · {categoria}",
            )
            for div_code in MATERIAL_TO_DIVISIONS.get(material_es, ["hidrosanitario"]):
                await ensure_product_division(s, sku, div_id_by_code[div_code])
            ok += 1

        await s.commit()

    print(f"Seeded {ok} products successfully.")
    if skipped:
        print(f"Skipped {len(skipped)}: {skipped[:5]}...")


if __name__ == "__main__":
    asyncio.run(main())
